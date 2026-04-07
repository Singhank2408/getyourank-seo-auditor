"""
GetYouRank — SEO Auditor Tool
Login/Signup + Full Technical SEO Audit
Run: python app.py
"""

import os, time, threading, uuid, json, re, sqlite3, hashlib, secrets
from collections import defaultdict
from urllib.parse import urljoin, urlparse
from functools import wraps
from datetime import datetime

import requests
from bs4 import BeautifulSoup
from flask import (Flask, request, jsonify, send_file, Response,
                   redirect, url_for, session, g)

app  = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))
jobs = {}

TIMEOUT      = 12
CRAWL_DELAY  = 0.5
LINK_DELAY   = 0.15
MAX_LINKS    = 800

REQ_HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; GetYouRankBot/1.0)",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml",
}

# ══════════════════════════════════════════════════════════════════════
# DATABASE — PostgreSQL (Render) or SQLite (local)
# ══════════════════════════════════════════════════════════════════════
try:
    import psycopg2
    import psycopg2.extras
    HAS_PG = True
except ImportError:
    HAS_PG = False

def get_database_url():
    url = os.environ.get("DATABASE_URL", "")
    if url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql://", 1)
    return url

USE_PG = bool(os.environ.get("DATABASE_URL", "")) and HAS_PG

def get_db():
    if "db" not in g:
        if USE_PG:
            g.db = psycopg2.connect(get_database_url())
            g.db.autocommit = False
        else:
            g.db = sqlite3.connect("/tmp/users.db")
            g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop("db", None)
    if db:
        try: db.close()
        except: pass

def db_execute(sql, params=(), fetchone=False, fetchall=False, commit=False):
    """Unified query helper for both PG and SQLite."""
    db = get_db()
    if USE_PG:
        sql = sql.replace("?", "%s")
        sql = sql.replace("INTEGER PRIMARY KEY AUTOINCREMENT", "SERIAL PRIMARY KEY")
        sql = sql.replace("datetime('now')", "NOW()")
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    else:
        cur = db.cursor()
    cur.execute(sql, params)
    result = None
    if fetchone:  result = cur.fetchone()
    if fetchall:  result = cur.fetchall()
    if commit:
        db.commit()
    return result

def init_db():
    with app.app_context():
        db = get_db()
        if USE_PG:
            cur = db.cursor()
            cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id       SERIAL PRIMARY KEY,
                    name     TEXT NOT NULL,
                    email    TEXT NOT NULL UNIQUE,
                    password TEXT NOT NULL,
                    plan     TEXT DEFAULT 'free',
                    created  TIMESTAMP DEFAULT NOW()
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS audits (
                    id         SERIAL PRIMARY KEY,
                    user_id    INTEGER NOT NULL,
                    site_name  TEXT,
                    url_count  INTEGER,
                    issues     INTEGER,
                    created    TIMESTAMP DEFAULT NOW()
                )
            """)
            db.commit()
        else:
            cur = db.cursor()
            cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id       INTEGER PRIMARY KEY AUTOINCREMENT,
                    name     TEXT NOT NULL,
                    email    TEXT NOT NULL UNIQUE,
                    password TEXT NOT NULL,
                    plan     TEXT DEFAULT 'free',
                    created  TEXT DEFAULT (datetime('now'))
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS audits (
                    id         INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id    INTEGER NOT NULL,
                    site_name  TEXT,
                    url_count  INTEGER,
                    issues     INTEGER,
                    created    TEXT DEFAULT (datetime('now'))
                )
            """)
            db.commit()

def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def get_user(email):
    db = get_db()
    if USE_PG:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM users WHERE email=%s", (email,))
        return cur.fetchone()
    else:
        cur = db.execute("SELECT * FROM users WHERE email=?", (email,))
        return cur.fetchone()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated

# ══════════════════════════════════════════════════════════════════════
# EXCEL HELPERS
# ══════════════════════════════════════════════════════════════════════
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def mkfill(h): return PatternFill("solid", start_color=h, end_color=h)
F = {
    "red": mkfill("FFCCCC"), "yellow": mkfill("FFF2CC"),
    "green": mkfill("D9EAD3"), "orange": mkfill("FCE5CD"),
    "hdr": mkfill("0A0A0A"), "alt": mkfill("F8FAFC"),
    "good": mkfill("D1FAE5"), "warn": mkfill("FEF3C7"), "bad": mkfill("FFE4E6"),
}
THIN = Border(
    left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),  bottom=Side(style="thin", color="E2E8F0"),
)
HF = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BF = Font(name="Calibri", size=9)

def hdr(ws, r, c, v, w=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = HF; cell.fill = F["hdr"]; cell.border = THIN
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if w: ws.column_dimensions[get_column_letter(c)].width = w
    return cell

def dat(ws, r, c, v, fk=None, wrap=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = BF; cell.border = THIN
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    if fk: cell.fill = F[fk]
    return cell

# ══════════════════════════════════════════════════════════════════════
# SITEMAP
# ══════════════════════════════════════════════════════════════════════
def parse_sitemap(url, max_urls=500):
    urls = []
    try:
        r = requests.get(url, headers=REQ_HEADERS, timeout=TIMEOUT)
        soup = BeautifulSoup(r.text, "lxml-xml")
        for loc in soup.find_all("sitemap"):
            child = loc.find("loc")
            if child and len(urls) < max_urls:
                urls.extend(parse_sitemap(child.text.strip(), max_urls - len(urls)))
        for loc in soup.find_all("url"):
            l = loc.find("loc")
            if l and len(urls) < max_urls:
                urls.append(l.text.strip())
    except Exception:
        pass
    return urls

# ══════════════════════════════════════════════════════════════════════
# FETCH + EXTRACT
# ══════════════════════════════════════════════════════════════════════
def fetch(url):
    try:
        t0 = time.time()
        r  = requests.get(url, headers=REQ_HEADERS, timeout=TIMEOUT, allow_redirects=True)
        ms = int((time.time() - t0) * 1000)
        return r.status_code, r.text, r.url, r.history, ms, len(r.content), None
    except requests.exceptions.Timeout:
        return None, None, url, [], 0, 0, "Timeout"
    except Exception as e:
        return None, None, url, [], 0, 0, str(e)[:80]

def extract_meta(html, base_url):
    soup = BeautifulSoup(html, "lxml")
    d    = {}
    t    = soup.find("title")
    d["title"]         = t.get_text(strip=True) if t else ""
    d["title_len"]     = len(d["title"])
    m = soup.find("meta", attrs={"name": re.compile(r"^description$", re.I)})
    d["meta_desc"]     = m.get("content","").strip() if m else ""
    d["meta_desc_len"] = len(d["meta_desc"])
    h1s = soup.find_all("h1")
    d["h1"]       = h1s[0].get_text(strip=True) if h1s else ""
    d["h1_count"] = len(h1s)
    d["h2_count"] = len(soup.find_all("h2"))
    c  = soup.find("link", attrs={"rel": lambda x: x and "canonical" in x})
    d["canonical"] = c.get("href","").strip() if c else ""
    rb = soup.find("meta", attrs={"name": re.compile(r"^robots$", re.I)})
    d["robots"]    = rb.get("content","").strip().lower() if rb else ""
    for prop in ["og:title","og:description","og:image","og:url"]:
        tag = soup.find("meta", attrs={"property": prop})
        d[prop.replace(":","_")] = tag.get("content","").strip() if tag else ""
    tc = soup.find("meta", attrs={"name": re.compile(r"^twitter:card$", re.I)})
    d["twitter_card"] = tc.get("content","").strip() if tc else ""
    schemas = soup.find_all("script", attrs={"type":"application/ld+json"})
    d["schema_count"] = len(schemas)
    types = []
    for s in schemas:
        try:
            obj = json.loads(s.string or "")
            tp  = obj.get("@type") or (obj.get("@graph") or [{}])[0].get("@type","")
            if tp: types.append(str(tp))
        except: pass
    d["schema_types"] = ", ".join(types)
    body = soup.find("body")
    d["word_count"] = len(body.get_text().split()) if body else 0
    return d, soup

def analyze_redirects(orig, history, final, status):
    hops  = len(history)
    chain = " → ".join([f"{r.url}[{r.status_code}]" for r in history] + [f"{final}[{status}]"])
    issues = []
    if hops > 0:
        if orig.startswith("http://") and final.startswith("https://"): issues.append("HTTP→HTTPS")
        op, fp = urlparse(orig), urlparse(final)
        if op.netloc.replace("www.","") != fp.netloc.replace("www.",""): issues.append("Domain change")
        if hops >= 3: issues.append(f"Long chain ({hops} hops)")
        elif hops == 2: issues.append("2-hop chain")
        else: issues.append("1 redirect")
    return {
        "redirect_hops": hops, "redirect_chain": chain,
        "redirect_issues": "; ".join(issues) if issues else "",
        "has_redirect": hops > 0,
    }

def extract_images(soup, page_url):
    imgs = []
    generic = {"image","photo","picture","img","icon","logo","banner","thumbnail"}
    for img in soup.find_all("img"):
        src = img.get("src","").strip()
        if not src or src.startswith("data:"): continue
        alt = img.get("alt", None)
        if alt is None:                             st = "Missing alt"
        elif alt.strip() == "":                     st = "Empty alt"
        elif alt.strip().lower() in generic:        st = f"Generic alt: '{alt}'"
        else:                                       st = "OK"
        imgs.append({
            "page_url": page_url,
            "img_src": urljoin(page_url, src),
            "alt_text": alt if alt is not None else "",
            "alt_status": st,
            "loading": img.get("loading",""),
            "has_issue": st != "OK",
        })
    return imgs

def extract_links(soup, page_url, site_domain):
    links = []
    site_domain = site_domain.replace("www.","")
    for a in soup.find_all("a", href=True):
        href = a.get("href","").strip()
        if not href or href.startswith(("#","mailto:","tel:","javascript:")): continue
        abs_url   = urljoin(page_url, href)
        parsed    = urlparse(abs_url)
        if not parsed.scheme.startswith("http"): continue
        link_dom  = parsed.netloc.replace("www.","")
        links.append({
            "source":   page_url,
            "url":      abs_url,
            "text":     a.get_text(strip=True)[:80],
            "type":     "internal" if link_dom == site_domain else "external",
            "nofollow": "nofollow" in (a.get("rel") or []),
        })
    return links

def meta_flags(row):
    f = []
    if not row.get("title"):              f.append("❌ Missing title")
    elif row["title_len"] > 60:           f.append(f"⚠ Title too long ({row['title_len']})")
    elif row["title_len"] < 30:           f.append(f"⚠ Title too short ({row['title_len']})")
    if not row.get("meta_desc"):          f.append("❌ Missing meta description")
    elif row["meta_desc_len"] > 160:      f.append(f"⚠ Meta desc too long ({row['meta_desc_len']})")
    elif row["meta_desc_len"] < 70:       f.append(f"⚠ Meta desc too short ({row['meta_desc_len']})")
    if not row.get("h1"):                 f.append("❌ Missing H1")
    elif row["h1_count"] > 1:             f.append(f"⚠ Multiple H1s ({row['h1_count']})")
    if not row.get("canonical"):          f.append("❌ Missing canonical")
    if "noindex" in row.get("robots",""): f.append("🚫 NOINDEX detected")
    if not row.get("og_title"):           f.append("⚠ Missing og:title")
    if not row.get("og_image"):           f.append("⚠ Missing og:image")
    if row.get("schema_count",0) == 0:    f.append("⚠ No structured data")
    return f

def find_dups(results, field):
    seen = defaultdict(list)
    for i, r in enumerate(results):
        v = str(r.get(field,"")).strip()
        if v: seen[v].append(i)
    return {idx for idxs in seen.values() if len(idxs) > 1 for idx in idxs}

# ══════════════════════════════════════════════════════════════════════
# PAGE SPEED
# ══════════════════════════════════════════════════════════════════════
PSI_API_KEY = os.environ.get("PSI_API_KEY", "")

def get_pagespeed(url, api_key=""):
    key = api_key or PSI_API_KEY
    params = {"url": url, "strategy": "mobile"}
    if key: params["key"] = key
    try:
        r    = requests.get("https://www.googleapis.com/pagespeedonline/v5/runPagespeed",
                            params=params, timeout=30)
        if r.status_code != 200: return {"error": f"HTTP {r.status_code}", "perf_score": "—"}
        data = r.json()
        lhr  = data.get("lighthouseResult",{})
        cats = lhr.get("categories",{})
        aud  = lhr.get("audits",{})
        def sl(s): return "Good" if s and s>=0.9 else ("Needs Improvement" if s and s>=0.5 else "Poor")
        def av(k): return aud.get(k,{}).get("displayValue","—")
        def asc(k): return sl(aud.get(k,{}).get("score"))
        return {
            "perf_score": int((cats.get("performance",{}).get("score") or 0)*100),
            "lcp": av("largest-contentful-paint"), "lcp_score": asc("largest-contentful-paint"),
            "cls": av("cumulative-layout-shift"),  "cls_score": asc("cumulative-layout-shift"),
            "tbt": av("total-blocking-time"),       "tbt_score": asc("total-blocking-time"),
            "fcp": av("first-contentful-paint"),    "ttfb": av("server-response-time"),
            "error": None,
        }
    except Exception as e:
        return {"error": str(e)[:60], "perf_score": "—"}

def run_pagespeed_batch(job_id, api_key):
    job  = jobs.get(job_id)
    if not job: return
    urls = [r["url"] for r in job.get("results",[]) if not r.get("fetch_error")][:20]
    job["speed_status"] = "running"; job["speed_total"] = len(urls); job["speed_progress"] = 0
    speed_map = {}
    for i, url in enumerate(urls):
        job["speed_progress"] = i+1
        speed_map[url] = get_pagespeed(url, api_key)
        time.sleep(1.2)
    job["speed_map"]    = speed_map
    job["speed_status"] = "done"

# ══════════════════════════════════════════════════════════════════════
# LINK CHECKER
# ══════════════════════════════════════════════════════════════════════
def check_link(url):
    try:
        r = requests.head(url, headers=REQ_HEADERS, timeout=8, allow_redirects=True)
        if r.status_code == 405:
            r = requests.get(url, headers=REQ_HEADERS, timeout=8, allow_redirects=True, stream=True); r.close()
        return r.status_code
    except: return "Error"

def run_link_checker(job_id):
    job = jobs.get(job_id)
    if not job: return
    seen   = {}
    for lnk in job.get("all_links",[]):
        if lnk["url"] not in seen: seen[lnk["url"]] = lnk
    unique = list(seen.keys())[:MAX_LINKS]
    job["link_total"] = len(unique); job["link_progress"] = 0; job["link_status"] = "running"
    statuses = {}
    for i, url in enumerate(unique):
        job["link_progress"] = i+1
        statuses[url] = check_link(url)
        time.sleep(LINK_DELAY)
    broken = []
    for lnk in job.get("all_links",[]):
        st = statuses.get(lnk["url"])
        if st and (str(st).startswith(("4","5")) or st == "Error"):
            broken.append({**lnk, "status": st})
    job["broken_links"]  = broken
    job["link_status"]   = "done"
    job["link_summary"]  = {
        "total_unique": len(unique), "broken": len(broken),
        "internal_broken": sum(1 for b in broken if b["type"]=="internal"),
        "external_broken": sum(1 for b in broken if b["type"]=="external"),
    }

# ══════════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ══════════════════════════════════════════════════════════════════════
META_COLS = [
    ("URL","url",52),("Status","status",8),("Resp ms","resp_ms",9),("Size KB","size_kb",8),
    ("Title","title",36),("Title Len","title_len",9),("Meta Desc","meta_desc",44),("Desc Len","meta_desc_len",9),
    ("H1","h1",30),("H1 Count","h1_count",8),("H2 Count","h2_count",8),
    ("Canonical","canonical",36),("Robots","robots",14),
    ("OG Title","og_title",30),("OG Image","og_image",30),
    ("Twitter","twitter_card",14),("Schema Types","schema_types",24),("Schema Count","schema_count",9),
    ("Word Count","word_count",9),("Issues","issues_text",55),
]

def build_excel(job, site_name):
    wb      = Workbook()
    results = job.get("results",[])
    broken  = job.get("broken_links",[])
    imgs    = job.get("all_images",[])
    speed   = job.get("speed_map",{})

    dup_sets = {f: find_dups(results, f) for f in ("title","meta_desc","h1","canonical","og_title")}

    total    = len(results)
    n_issues = sum(1 for r in results if r.get("flags"))
    n_redir  = sum(1 for r in results if r.get("has_redirect"))
    n_noidx  = sum(1 for r in results if "noindex" in str(r.get("robots","")).lower())

    # ── Sheet 1 ───────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Full Audit"; ws.freeze_panes = "A3"
    ws.merge_cells("A1:E1"); ws.merge_cells("F1:J1")
    ws.merge_cells("K1:O1"); ws.merge_cells("P1:T1")
    for val, addr, color in [
        (f"Total: {total}","A1","1E293B"), (f"Issues: {n_issues}","F1","991B1B"),
        (f"Redirects: {n_redir}","K1","92400E"), (f"NOINDEX: {n_noidx}","P1","991B1B"),
    ]:
        c = ws[addr]; c.value = val
        c.font = Font(name="Calibri", bold=True, size=10, color=color)
        c.fill = mkfill("F1F5F9"); c.border = THIN
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for ci, (label, _, w) in enumerate(META_COLS, 1): hdr(ws, 2, ci, label, w)
    ws.row_dimensions[2].height = 30

    WRAP_CI = {1,5,7,9,12,14,15,17,20}
    for ri, row in enumerate(results, 3):
        alt = ri % 2 == 0
        for ci, (_, field, _) in enumerate(META_COLS, 1):
            val = row.get(field,"")
            c   = dat(ws, ri, ci, val, "alt" if alt else None, ci in WRAP_CI)
            if row.get("fetch_error"):                c.fill = F["orange"]; continue
            if field in ("title","meta_desc","h1","canonical","og_title","og_image") and not str(val).strip():
                c.fill = F["red"]; c.font = Font(name="Calibri", size=9, color="7F1D1D"); continue
            if field in dup_sets and (ri-3) in dup_sets[field]:
                c.fill = F["yellow"]; c.font = Font(name="Calibri", size=9, color="78350F"); continue
            if field == "robots" and "noindex" in str(val).lower():
                c.fill = F["red"]; c.font = Font(name="Calibri", size=9, bold=True, color="7F1D1D"); continue
            if field in ("title_len","meta_desc_len") and val:
                lim = (30,60) if field=="title_len" else (70,160)
                if not (lim[0]<=int(val)<=lim[1]): c.fill = F["yellow"]
            if field == "schema_count" and int(val or 0)==0: c.fill = F["yellow"]
            if field == "issues_text" and str(val).strip():
                c.fill = F["red"]; c.font = Font(name="Calibri", size=9, italic=True, color="475569")
    ws.auto_filter.ref = f"A2:{get_column_letter(len(META_COLS))}2"

    # ── Sheet 2: Issues ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Issues Summary"); ws2.freeze_panes = "A2"
    for ci,(h,w) in enumerate([("URL",55),("Status",9),("Resp ms",9),("Count",9),("Issues",80)],1):
        hdr(ws2, 1, ci, h, w)
    for ri,(url,st,ms,flags) in enumerate(
        sorted([(r["url"],r.get("status",""),r.get("resp_ms",""),r.get("flags",[])) for r in results if r.get("flags")],
               key=lambda x:len(x[3]),reverse=True), 2):
        fk = "bad" if len(flags)>=4 else ("warn" if len(flags)>=2 else "yellow")
        for ci,val in enumerate([url,st,ms,len(flags)," | ".join(flags)],1):
            dat(ws2, ri, ci, val, fk, ci in {1,5})

    # ── Sheet 3: Redirects ────────────────────────────────────────────
    ws3 = wb.create_sheet("Redirect Report"); ws3.freeze_panes = "A2"
    for ci,(h,w) in enumerate([("Original URL",50),("Final URL",50),("Hops",7),("Issues",30),("Chain",80)],1):
        hdr(ws3, 1, ci, h, w)
    for ri,row in enumerate(sorted(results, key=lambda x:x.get("redirect_hops",0),reverse=True), 2):
        hops = row.get("redirect_hops",0)
        fk   = "bad" if hops>=3 else ("warn" if hops>=2 else ("yellow" if hops==1 else None))
        for ci,val in enumerate([row["url"],row.get("final_url",""),hops,row.get("redirect_issues",""),row.get("redirect_chain","")],1):
            dat(ws3, ri, ci, val, fk, ci in {1,2,5})

    # ── Sheet 4: Broken Links ─────────────────────────────────────────
    ws4 = wb.create_sheet("Broken Links"); ws4.freeze_panes = "A2"
    for ci,(h,w) in enumerate([("Source Page",52),("Broken URL",52),("Status",9),("Type",11),("Link Text",35)],1):
        hdr(ws4, 1, ci, h, w)
    if broken:
        for ri,lnk in enumerate(broken, 2):
            st  = str(lnk.get("status",""))
            fk  = "bad" if st.startswith("4") or st=="Error" else "warn"
            for ci,val in enumerate([lnk["source"],lnk["url"],lnk.get("status",""),lnk.get("type",""),lnk.get("text","")],1):
                dat(ws4, ri, ci, val, fk, ci in {1,2,5})
    else:
        ws4.cell(row=2,column=1,value="No broken links found!").font = Font(name="Calibri",italic=True,color="16A34A")

    # ── Sheet 5: Images ───────────────────────────────────────────────
    ws5 = wb.create_sheet("Image Audit"); ws5.freeze_panes = "A2"
    for ci,(h,w) in enumerate([("Page URL",50),("Image URL",50),("Alt Text",35),("Alt Status",16),("Loading",10)],1):
        hdr(ws5, 1, ci, h, w)
    for ri,img in enumerate(sorted(imgs, key=lambda x:x.get("has_issue",False),reverse=True), 2):
        st  = img.get("alt_status","OK")
        fk  = "red" if "Missing" in st else ("yellow" if st!="OK" else None)
        for ci,val in enumerate([img["page_url"],img["img_src"],img.get("alt_text",""),st,img.get("loading","")],1):
            dat(ws5, ri, ci, val, fk, ci in {1,2,3})

    # ── Sheet 6: Page Speed ───────────────────────────────────────────
    ws6 = wb.create_sheet("Page Speed (CWV)"); ws6.freeze_panes = "A2"
    for ci,(h,w) in enumerate([("URL",52),("Score",10),("LCP",14),("LCP Rating",14),("CLS",12),("CLS Rating",14),("TBT",14),("FCP",14),("TTFB",14)],1):
        hdr(ws6, 1, ci, h, w)
    if speed:
        for ri,(url,sp) in enumerate(speed.items(), 2):
            score = sp.get("perf_score","—")
            sfk   = "good" if isinstance(score,int) and score>=90 else ("warn" if isinstance(score,int) and score>=50 else "bad")
            def rfk(v): return {"Good":"good","Needs Improvement":"warn","Poor":"bad"}.get(v)
            vals = [url,score,sp.get("lcp","—"),sp.get("lcp_score","—"),sp.get("cls","—"),sp.get("cls_score","—"),sp.get("tbt","—"),sp.get("fcp","—"),sp.get("ttfb","—")]
            for ci,val in enumerate(vals, 1):
                dat(ws6, ri, ci, val, sfk if ci==2 else (rfk(val) if ci in {4,6} else None), ci==1)
    else:
        ws6.cell(row=2,column=1,value="Run Speed Audit from dashboard to populate this sheet.").font = Font(name="Calibri",italic=True,color="64748B")

    # ── Sheet 7: Duplicates ───────────────────────────────────────────
    ws7 = wb.create_sheet("Duplicate Report"); ws7.freeze_panes = "A2"
    for ci,(h,w) in enumerate([("Field",18),("Value",50),("Affected URLs",80),("Count",8)],1):
        hdr(ws7, 1, ci, h, w)
    dup_row = 2
    for label,field in [("Title","title"),("Meta Desc","meta_desc"),("H1","h1"),("Canonical","canonical"),("OG Title","og_title")]:
        seen2 = defaultdict(list)
        for r in results:
            v = str(r.get(field,"")).strip()
            if v: seen2[v].append(r["url"])
        for v,urls in seen2.items():
            if len(urls)>1:
                for ci,val in enumerate([label,v,"\n".join(urls),len(urls)],1):
                    dat(ws7, dup_row, ci, val, "yellow", True)
                dup_row += 1
    if dup_row==2:
        ws7.cell(row=2,column=1,value="No duplicates found.").font = Font(name="Calibri",italic=True,color="16A34A")

    fname = f"getyourank_audit_{int(time.time())}.xlsx"
    path  = f"/tmp/{fname}"
    wb.save(path)
    return path, fname

# ══════════════════════════════════════════════════════════════════════
# BACKGROUND WORKER
# ══════════════════════════════════════════════════════════════════════
def run_audit(job_id, urls, site_name):
    job = jobs[job_id]
    job["total"]     = len(urls)
    job["all_links"] = []
    job["all_images"]= []
    site_domain = urlparse(urls[0]).netloc if urls else ""
    results = []

    for i, url in enumerate(urls):
        if job.get("cancelled"): break
        job["progress"] = i; job["current_url"] = url
        status, html, final_url, history, ms, size, error = fetch(url)
        row = {"url": url}
        if error or not html:
            row.update({
                "status":error or "No response","fetch_error":True,"resp_ms":0,"size_kb":0,
                "title":"","title_len":0,"meta_desc":"","meta_desc_len":0,"h1":"","h1_count":0,
                "h2_count":0,"canonical":"","robots":"","og_title":"","og_description":"",
                "og_image":"","og_url":"","twitter_card":"","schema_count":0,"schema_types":"",
                "word_count":0,"redirect_hops":0,"redirect_chain":"","redirect_issues":"",
                "has_redirect":False,"final_url":url,
                "flags":[f"❌ Fetch failed: {error}"],"issues_text":f"Fetch failed: {error}",
            })
        else:
            meta, soup = extract_meta(html, final_url)
            redir      = analyze_redirects(url, history, final_url, status)
            job["all_images"].extend(extract_images(soup, final_url))
            job["all_links"].extend(extract_links(soup, final_url, site_domain))
            flags = meta_flags(meta)
            row.update(meta); row.update(redir)
            row.update({
                "url":url,"final_url":final_url,"status":status,"fetch_error":False,
                "resp_ms":ms,"size_kb":round(size/1024,1),"flags":flags,"issues_text":" | ".join(flags),
            })
        results.append(row)
        job["results"] = results
        job["log"].append({"url":url,"status":row.get("status",""),"ms":row.get("resp_ms",0),"issues":len(row.get("flags",[]))})
        time.sleep(CRAWL_DELAY)

    job["phase"] = "link_check"
    threading.Thread(target=run_link_checker, args=(job_id,), daemon=True).start()
    for _ in range(120):
        if job.get("link_status") == "done": break
        time.sleep(0.5)

    try:
        path, fname = build_excel(job, site_name)
        job["excel_path"] = path; job["excel_name"] = fname
    except Exception as e:
        job["excel_error"] = str(e)

    # Save audit record
    try:
        with app.app_context():
            db = get_db()
            if USE_PG:
                cur = db.cursor()
                cur.execute("INSERT INTO audits(user_id,site_name,url_count,issues) VALUES(%s,%s,%s,%s)",
                           (job["user_id"], site_name, len(results), sum(1 for r in results if r.get("flags"))))
                db.commit()
            else:
                db.execute("INSERT INTO audits(user_id,site_name,url_count,issues) VALUES(?,?,?,?)",
                           (job["user_id"], site_name, len(results), sum(1 for r in results if r.get("flags"))))
                db.commit()
    except: pass

    job["status"] = "done"; job["progress"] = len(results)

# ══════════════════════════════════════════════════════════════════════
# AUTH ROUTES
# ══════════════════════════════════════════════════════════════════════
@app.route("/")
def home():
    if "user_id" in session: return redirect(url_for("dashboard"))
    return redirect(url_for("login_page"))

@app.route("/login", methods=["GET"])
def login_page():
    if "user_id" in session: return redirect(url_for("dashboard"))
    return Response(render_auth("login"), mimetype="text/html")

@app.route("/signup", methods=["GET"])
def signup_page():
    if "user_id" in session: return redirect(url_for("dashboard"))
    return Response(render_auth("signup"), mimetype="text/html")

@app.route("/api/login", methods=["POST"])
def do_login():
    try:
        data  = request.json
        email = data.get("email","").strip().lower()
        pw    = data.get("password","")
        user  = get_user(email)
        if not user or user["password"] != hash_pw(pw):
            return jsonify({"error": "Invalid email or password"}), 401
        session["user_id"]   = user["id"]
        session["user_name"] = user["name"]
        session["user_email"]= user["email"]
        return jsonify({"ok": True})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Login failed: {str(e)}"}), 500

@app.route("/api/signup", methods=["POST"])
def do_signup():
    try:
        data  = request.json
        name  = data.get("name","").strip()
        email = data.get("email","").strip().lower()
        pw    = data.get("password","")
        if not name or not email or not pw:
            return jsonify({"error": "All fields are required"}), 400
        if len(pw) < 6:
            return jsonify({"error": "Password must be at least 6 characters"}), 400
        db = get_db()
        if USE_PG:
            cur = db.cursor()
            cur.execute("SELECT id FROM users WHERE email=%s", (email,))
            if cur.fetchone():
                return jsonify({"error": "Email already registered"}), 400
            cur.execute("INSERT INTO users(name,email,password) VALUES(%s,%s,%s)", (name, email, hash_pw(pw)))
            db.commit()
        else:
            if db.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone():
                return jsonify({"error": "Email already registered"}), 400
            db.execute("INSERT INTO users(name,email,password) VALUES(?,?,?)", (name, email, hash_pw(pw)))
            db.commit()
        user = get_user(email)
        if not user:
            return jsonify({"error": "Account created but login failed — please try signing in"}), 500
        session["user_id"]   = user["id"]
        session["user_name"] = user["name"]
        session["user_email"]= user["email"]
        return jsonify({"ok": True})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Signup failed: {str(e)}"}), 500

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))

# ══════════════════════════════════════════════════════════════════════
# DASHBOARD + AUDIT ROUTES
# ══════════════════════════════════════════════════════════════════════
@app.route("/dashboard")
@login_required
def dashboard():
    try:
        db = get_db()
        if USE_PG:
            cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute("SELECT * FROM audits WHERE user_id=%s ORDER BY created DESC LIMIT 10", (session["user_id"],))
            audits = cur.fetchall()
        else:
            audits = db.execute("SELECT * FROM audits WHERE user_id=? ORDER BY created DESC LIMIT 10",
                                (session["user_id"],)).fetchall()
    except Exception as e:
        print(f"Dashboard DB error: {e}")
        audits = []
    try:
        html = render_dashboard(audits)
        return Response(html, mimetype="text/html")
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"Dashboard render error: {str(e)}", 500

@app.route("/api/start", methods=["POST"])
@login_required
def start_audit():
    data       = request.json
    site_name  = data.get("site_name","Website").strip() or "Website"
    max_urls   = min(int(data.get("max_urls",100)), 10000)
    sitemap_url= data.get("sitemap_url","").strip()
    raw        = data.get("urls","")

    MAX_URLS_PER_USER = 10
    if sitemap_url:
        urls = parse_sitemap(sitemap_url, MAX_URLS_PER_USER)
        if not urls: return jsonify({"error": "Could not parse sitemap."}), 400
    else:
        urls = [u.strip() for u in raw.splitlines() if u.strip().startswith("http")]
        urls = list(dict.fromkeys(urls))[:MAX_URLS_PER_USER]
    if not urls: return jsonify({"error": "No valid URLs found."}), 400
    if len(urls) > MAX_URLS_PER_USER:
        urls = urls[:MAX_URLS_PER_USER]

    job_id = str(uuid.uuid4())[:8]
    jobs[job_id] = {
        "status":"running","progress":0,"total":len(urls),
        "results":[],"log":[],"current_url":"","phase":"crawl",
        "all_links":[],"all_images":[],"broken_links":[],
        "link_status":"pending","link_progress":0,"link_total":0,
        "speed_status":"idle","speed_map":{},
        "user_id": session["user_id"],
    }
    threading.Thread(target=run_audit, args=(job_id,urls,site_name), daemon=True).start()
    return jsonify({"job_id": job_id, "total": len(urls)})

@app.route("/api/status/<job_id>")
@login_required
def job_status(job_id):
    job = jobs.get(job_id)
    if not job: return jsonify({"error":"Not found"}), 404
    results = job.get("results",[])
    broken  = job.get("broken_links",[])
    imgs    = job.get("all_images",[])
    return jsonify({
        "status":job["status"],"phase":job.get("phase","crawl"),
        "progress":job["progress"],"total":job["total"],
        "current_url":job.get("current_url",""),
        "log":job["log"][-30:],
        "has_excel":bool(job.get("excel_path")),
        "link_status":job.get("link_status","pending"),
        "link_progress":job.get("link_progress",0),"link_total":job.get("link_total",0),
        "speed_status":job.get("speed_status","idle"),
        "speed_progress":job.get("speed_progress",0),"speed_total":job.get("speed_total",0),
        "summary":{
            "total":len(results),"issues":sum(1 for r in results if r.get("flags")),
            "fetch_errors":sum(1 for r in results if r.get("fetch_error")),
            "noindex":sum(1 for r in results if "noindex" in str(r.get("robots","")).lower()),
            "redirects":sum(1 for r in results if r.get("has_redirect")),
            "missing_title":sum(1 for r in results if not r.get("title")),
            "missing_desc":sum(1 for r in results if not r.get("meta_desc")),
            "missing_h1":sum(1 for r in results if not r.get("h1")),
            "missing_canonical":sum(1 for r in results if not r.get("canonical")),
            "no_schema":sum(1 for r in results if r.get("schema_count",0)==0),
            "slow_pages":sum(1 for r in results if int(r.get("resp_ms",0))>2000),
            "broken_links":len(broken),
            "internal_broken":sum(1 for b in broken if b.get("type")=="internal"),
            "total_images":len(imgs),
            "img_issues":sum(1 for i in imgs if i.get("has_issue")),
        }
    })

@app.route("/api/psi-key")
@login_required
def get_psi_key():
    """Return PSI API key to authenticated users only — browser makes calls directly."""
    key = PSI_API_KEY
    if not key:
        return jsonify({"error": "No API key configured on server"}), 404
    return jsonify({"key": key})

@app.route("/api/save-speed/<job_id>", methods=["POST"])
@login_required
def save_speed_results(job_id):
    """Receive speed results from browser and save to job for Excel rebuild."""
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found — please re-run the audit first"}), 404
    results = request.json.get("results", {})
    job["speed_map"] = results
    # Rebuild Excel with speed data
    try:
        path, fname = build_excel(job, job.get("site_name", "Website"))
        job["excel_path"] = path
        job["excel_name"] = fname
        return jsonify({"ok": True, "message": f"Excel rebuilt with speed data for {len(results)} URLs"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/speed/direct", methods=["POST"])
@login_required
def run_speed_direct():
    """Synchronous speed check — returns results directly, no background thread needed."""
    data    = request.json
    urls    = data.get("urls", [])[:5]
    if not urls:
        return jsonify({"error": "No URLs provided"}), 400
    results = {}
    for url in urls:
        results[url] = get_pagespeed(url, "")
        time.sleep(0.5)
    return jsonify({"results": results})

@app.route("/api/speed/<job_id>", methods=["POST"])
@login_required
def run_speed(job_id):
    job = jobs.get(job_id)
    if not job: return jsonify({"error":"Not found"}), 404
    api_key = request.json.get("api_key","")
    threading.Thread(target=run_pagespeed_batch, args=(job_id,""), daemon=True).start()
    return jsonify({"started":True})

@app.route("/api/download/<job_id>")
@login_required
def download(job_id):
    job = jobs.get(job_id)
    if not job or not job.get("excel_path"): return jsonify({"error":"Not ready"}), 404
    return send_file(job["excel_path"], as_attachment=True,
                     download_name=job.get("excel_name","seo_audit.xlsx"))

# ══════════════════════════════════════════════════════════════════════
# HTML TEMPLATES
# ══════════════════════════════════════════════════════════════════════
AUTH_STYLE = """
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'DM Sans',sans-serif;background:#0a0a0a;color:#f0f0f0;min-height:100vh;display:flex;flex-direction:column}
.top{padding:24px 40px;border-bottom:1px solid #1a1a1a}
.logo{font-size:18px;font-weight:700;letter-spacing:-.5px;color:#fff;text-decoration:none}
.logo span{color:#22c55e}
.auth-wrap{flex:1;display:flex;align-items:center;justify-content:center;padding:40px 20px}
.auth-box{width:100%;max-width:420px}
.auth-eyebrow{font-size:11px;font-weight:600;letter-spacing:.15em;text-transform:uppercase;color:#22c55e;margin-bottom:12px}
.auth-title{font-size:32px;font-weight:800;letter-spacing:-1px;line-height:1.1;margin-bottom:8px;text-transform:uppercase}
.auth-sub{font-size:13px;color:#666;margin-bottom:32px;line-height:1.6}
.field{display:flex;flex-direction:column;gap:6px;margin-bottom:16px}
.field label{font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.08em;color:#888}
.field input{background:#111;border:1px solid #222;border-radius:8px;color:#f0f0f0;font-size:14px;padding:12px 16px;outline:none;transition:border .15s;width:100%}
.field input:focus{border-color:#22c55e}
.err{background:#1a0808;border:1px solid #3f1f1f;border-radius:8px;color:#f87171;font-size:13px;padding:10px 14px;margin-bottom:16px;display:none}
.submit-btn{width:100%;background:#22c55e;border:none;border-radius:8px;color:#0a0a0a;font-size:14px;font-weight:700;padding:14px;cursor:pointer;letter-spacing:.5px;text-transform:uppercase;transition:opacity .15s;margin-top:8px}
.submit-btn:hover{opacity:.9}
.submit-btn:disabled{opacity:.4;cursor:not-allowed}
.switch{text-align:center;margin-top:20px;font-size:13px;color:#666}
.switch a{color:#22c55e;text-decoration:none;font-weight:500}
.divider{display:flex;align-items:center;gap:12px;margin:20px 0;color:#333;font-size:12px}
.divider::before,.divider::after{content:'';flex:1;height:1px;background:#1a1a1a}
.benefits{background:#0f0f0f;border:1px solid #1a1a1a;border-radius:10px;padding:20px;margin-bottom:24px}
.benefit{display:flex;align-items:center;gap:10px;font-size:12px;color:#888;margin-bottom:8px}
.benefit:last-child{margin-bottom:0}
.benefit-icon{color:#22c55e;font-size:14px}
</style>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
"""

def render_auth(mode):
    is_login = mode == "login"
    title    = "WELCOME BACK." if is_login else "CREATE ACCOUNT."
    sub      = "Sign in to access your SEO audit dashboard." if is_login else "Free account. Unlimited audits. No credit card."
    btn      = "Sign In →" if is_login else "Create Account →"
    switch   = 'Don\'t have an account? <a href="/signup">Sign up free</a>' if is_login else 'Already have an account? <a href="/login">Sign in</a>'
    action   = "/api/login" if is_login else "/api/signup"

    name_field = "" if is_login else """
    <div class="field"><label>Full Name</label><input type="text" id="name" placeholder="Ankit Singh" required></div>
    """
    benefits = "" if is_login else """
    <div class="benefits">
      <div class="benefit"><span class="benefit-icon">✓</span> Full technical SEO audit (meta, schema, CWV)</div>
      <div class="benefit"><span class="benefit-icon">✓</span> Redirect chain & broken link detection</div>
      <div class="benefit"><span class="benefit-icon">✓</span> Image alt text audit</div>
      <div class="benefit"><span class="benefit-icon">✓</span> Downloadable Excel report (7 sheets)</div>
    </div>
    """
    return f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{'Sign In' if is_login else 'Sign Up'} — GetYouRank SEO Auditor</title>
{AUTH_STYLE}</head>
<body>
<div class="top"><a class="logo" href="https://getyourank.com">GET<span>YOU</span>RANK</a></div>
<div class="auth-wrap">
  <div class="auth-box">
    <div class="auth-eyebrow">SEO Audit Tool</div>
    <h1 class="auth-title">{title}</h1>
    <p class="auth-sub">{sub}</p>
    {benefits}
    <div class="err" id="err"></div>
    {name_field}
    <div class="field"><label>Email Address</label><input type="email" id="email" placeholder="you@company.com" required></div>
    <div class="field"><label>Password</label><input type="password" id="password" placeholder="{'Your password' if is_login else 'Min. 6 characters'}" required></div>
    <button class="submit-btn" id="btn" onclick="submit()">{btn}</button>
    <div class="switch">{switch}</div>
  </div>
</div>
<script>
async function submit(){{
  const btn=document.getElementById('btn');
  const err=document.getElementById('err');
  err.style.display='none'; btn.disabled=true; btn.textContent='Please wait...';
  const body={{'email':document.getElementById('email').value,'password':document.getElementById('password').value}};
  {'body.name=document.getElementById("name").value;' if not is_login else ''}
  const res=await fetch('{action}',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify(body)}});
  const data=await res.json();
  if(data.error){{err.textContent=data.error;err.style.display='block';btn.disabled=false;btn.textContent='{btn}';}}
  else window.location.href='/dashboard';
}}
document.addEventListener('keydown',e=>{{if(e.key==='Enter')submit();}});
</script>
</body></html>"""

DASH_STYLE = """
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'DM Sans',sans-serif;background:#0a0a0a;color:#f0f0f0;min-height:100vh;display:grid;grid-template-rows:60px 1fr}
nav{background:#0f0f0f;border-bottom:1px solid #1a1a1a;display:flex;align-items:center;justify-content:space-between;padding:0 32px}
.logo{font-size:16px;font-weight:700;letter-spacing:-.3px;color:#fff;text-decoration:none}
.logo span{color:#22c55e}
.nav-right{display:flex;align-items:center;gap:20px}
.user-tag{font-size:12px;color:#555}
.nav-link{font-size:12px;color:#666;text-decoration:none;transition:color .15s}
.nav-link:hover{color:#22c55e}
.logout{font-size:12px;color:#444;text-decoration:none;border:1px solid #1a1a1a;padding:6px 12px;border-radius:6px;transition:all .15s}
.logout:hover{border-color:#333;color:#888}
main{display:grid;grid-template-columns:360px 1fr;overflow:hidden;height:calc(100vh - 60px)}
aside{background:#0f0f0f;border-right:1px solid #1a1a1a;padding:24px 20px;overflow-y:auto;display:flex;flex-direction:column;gap:18px}
.panel{overflow-y:auto;padding:24px 28px;display:flex;flex-direction:column;gap:20px}
.section-h{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.12em;color:#444;margin-bottom:10px}
.fg{display:flex;flex-direction:column;gap:6px}
.fg label{font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.08em;color:#555}
input[type=text],textarea{background:#0a0a0a;border:1px solid #1a1a1a;border-radius:7px;color:#f0f0f0;font-size:12px;padding:9px 12px;width:100%;outline:none;transition:border .15s}
input[type=text]:focus,textarea:focus{border-color:#22c55e}
textarea{resize:vertical;min-height:110px;font-family:'DM Mono',monospace;font-size:11px;line-height:1.6}
.tabs{display:flex;gap:2px;background:#0a0a0a;border-radius:7px;padding:2px;border:1px solid #1a1a1a}
.tab{flex:1;padding:6px;font-size:11px;font-weight:500;border:none;border-radius:5px;background:transparent;color:#555;cursor:pointer;transition:all .15s}
.tab.on{background:#22c55e;color:#0a0a0a;font-weight:700}
.slider-row{display:flex;align-items:center;gap:10px}
.slider-row input{flex:1;accent-color:#22c55e}
.sv{font-size:11px;font-family:'DM Mono',monospace;color:#22c55e;min-width:36px;text-align:right}
.run-btn{background:#22c55e;border:none;border-radius:8px;color:#0a0a0a;font-size:13px;font-weight:800;padding:13px;cursor:pointer;text-transform:uppercase;letter-spacing:.5px;width:100%;transition:opacity .15s}
.run-btn:hover{opacity:.88}
.run-btn:disabled{opacity:.3;cursor:not-allowed}
.hint{font-size:10px;color:#333;line-height:1.6}
.extra{display:none;flex-direction:column;gap:7px}
.extra.show{display:flex}
.toggle-btn{font-size:10px;color:#444;background:none;border:none;cursor:pointer;padding:0;display:flex;align-items:center;gap:4px}
.toggle-btn:hover{color:#888}
/* right panel */
.idle{display:flex;flex-direction:column;align-items:center;justify-content:center;min-height:50vh;gap:14px;text-align:center}
.idle-ico{font-size:44px;opacity:.15}
.idle-h{font-size:22px;font-weight:800;color:#f0f0f0;opacity:.3;text-transform:uppercase;letter-spacing:-1px}
.idle-sub{font-size:12px;color:#333;max-width:260px;line-height:1.7}
/* progress */
.prog-hdr{display:flex;justify-content:space-between;align-items:baseline;margin-bottom:8px}
.prog-t{font-size:13px;font-weight:600}
.prog-c{font-size:11px;font-family:'DM Mono',monospace;color:#444}
.prog-bg{background:#111;border-radius:99px;height:4px;overflow:hidden}
.prog-fill{height:100%;border-radius:99px;background:#22c55e;transition:width .3s ease}
.cur-url{font-size:10px;color:#333;font-family:'DM Mono',monospace;margin-top:6px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.badge{display:inline-block;font-size:9px;font-family:'DM Mono',monospace;padding:2px 8px;border-radius:4px;margin-top:6px}
.b-crawl{background:#0c1a2e;color:#60a5fa}
.b-links{background:#0a1a0f;color:#4ade80}
.b-done{background:#0a1a0f;color:#22c55e}
/* result tabs */
.rtab-bar{display:flex;gap:2px;background:#111;border-radius:8px;padding:3px;border:1px solid #1a1a1a}
.rtab{flex:1;padding:7px 4px;font-size:10px;font-weight:600;border:none;border-radius:5px;background:transparent;color:#444;cursor:pointer;transition:all .15s;white-space:nowrap;text-transform:uppercase;letter-spacing:.05em}
.rtab.on{background:#0f0f0f;color:#f0f0f0;border:1px solid #1a1a1a}
.rtab-c{display:none}
.rtab-c.on{display:block}
/* cards */
.grid4{display:grid;grid-template-columns:repeat(4,1fr);gap:10px}
.grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}
.grid2{display:grid;grid-template-columns:repeat(2,1fr);gap:10px}
.card{background:#0f0f0f;border:1px solid #1a1a1a;border-radius:8px;padding:14px}
.card-n{font-size:28px;font-weight:800;font-family:'DM Mono',monospace;line-height:1}
.card-l{font-size:9px;color:#444;margin-top:4px;text-transform:uppercase;letter-spacing:.08em;font-weight:600}
.card.red .card-n{color:#f87171}
.card.ora .card-n{color:#fb923c}
.card.grn .card-n{color:#22c55e}
.card.neu .card-n{color:#f0f0f0}
/* issue rows */
.irow{display:flex;justify-content:space-between;align-items:center;background:#0f0f0f;border:1px solid #1a1a1a;border-radius:7px;padding:9px 13px}
.irow-n{font-size:11px;color:#555}
.ib{font-size:13px;font-weight:700;font-family:'DM Mono',monospace}
.ib.r{color:#f87171}.ib.o{color:#fb923c}.ib.g{color:#22c55e}
/* log */
.log{background:#0f0f0f;border:1px solid #1a1a1a;border-radius:8px;padding:12px;max-height:260px;overflow-y:auto}
.lrow{display:flex;align-items:center;gap:8px;padding:4px 0;border-bottom:1px solid #111;font-size:10px;font-family:'DM Mono',monospace}
.lrow:last-child{border:none}
.ls{min-width:36px;font-weight:600}
.ls.ok{color:#22c55e}.ls.err{color:#f87171}.ls.rdr{color:#fb923c}
.lms{min-width:52px;text-align:right;color:#333}
.lu{flex:1;color:#444;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.li{color:#fb923c;min-width:18px;text-align:right}
/* history */
.history-row{display:flex;justify-content:space-between;align-items:center;padding:10px 0;border-bottom:1px solid #111;font-size:12px}
.history-row:last-child{border:none}
.h-site{font-weight:500}
.h-meta{font-size:10px;color:#444;font-family:'DM Mono',monospace}
.h-badge{font-size:10px;padding:2px 8px;border-radius:4px;background:#1a0a0a;color:#f87171}
/* download */
.dl-btn{background:#22c55e;border:none;border-radius:7px;color:#0a0a0a;font-size:12px;font-weight:700;padding:11px 22px;cursor:pointer;text-transform:uppercase;letter-spacing:.5px;transition:opacity .15s}
.dl-btn:hover{opacity:.88}
.dl-btn:disabled{opacity:.3;cursor:not-allowed;background:#333;color:#666}
.done-tag{font-size:10px;font-family:'DM Mono',monospace;border:1px solid #22c55e;color:#22c55e;padding:3px 10px;border-radius:5px;font-weight:600}
.speed-btn{background:transparent;border:1px solid #22c55e;border-radius:7px;color:#22c55e;font-size:11px;font-weight:600;padding:8px 16px;cursor:pointer;text-transform:uppercase;letter-spacing:.3px;transition:all .15s}
.speed-btn:hover{background:#22c55e;color:#0a0a0a}
.speed-btn:disabled{opacity:.3;cursor:not-allowed}
@media(max-width:860px){main{grid-template-columns:1fr}.grid4{grid-template-columns:repeat(2,1fr)}}
</style>
<link href="https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
"""

def render_dashboard(audits):
    history_rows = ""
    for a in audits:
        created = a["created"]
        if created and not isinstance(created, str):
            created = str(created)
        dt = created[:10] if created else ""
        history_rows += f"""
        <div class="history-row">
          <div><div class="h-site">{a['site_name'] or 'Unknown'}</div>
          <div class="h-meta">{dt} · {a['url_count']} URLs</div></div>
          <span class="h-badge">{a['issues']} issues</span>
        </div>"""
    if not history_rows:
        history_rows = '<div style="font-size:12px;color:#333;padding:10px 0">No audits yet. Run your first audit →</div>'

    return f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>SEO Auditor — GetYouRank</title>
{DASH_STYLE}</head>
<body>
<nav>
  <a class="logo" href="https://getyourank.com">GET<span>YOU</span>RANK</a>
  <div class="nav-right">
    <span class="user-tag">{session.get('user_name','')}</span>
    <a class="nav-link" href="https://getyourank.com" target="_blank">Back to site</a>
    <a class="logout" href="/logout">Sign out</a>
  </div>
</nav>
<main>
<aside>
  <div>
    <div class="section-h">New Audit</div>
    <div class="fg"><label>Site Name</label>
    <input type="text" id="siteName" placeholder="e.g. example.com"></div>
  </div>
  <div class="fg"><label>Input Method</label>
    <div style="display:flex;gap:3px;background:#0a0a0a;border-radius:7px;padding:3px;border:1px solid #1a1a1a">
      <button id="btn-urls" onclick="switchTab('urls')" style="flex:1;padding:6px;font-size:11px;font-weight:700;border:none;border-radius:5px;background:#22c55e;color:#0a0a0a;cursor:pointer;transition:all .15s">Paste URLs</button>
      <button id="btn-sitemap" onclick="switchTab('sitemap')" style="flex:1;padding:6px;font-size:11px;font-weight:500;border:none;border-radius:5px;background:transparent;color:#555;cursor:pointer;transition:all .15s">Sitemap</button>
    </div>
  </div>
  <div class="fg" id="t-urls">
    <label>URLs (one per line — max 10)</label>
    <textarea id="urlInput" placeholder="https://example.com/&#10;https://example.com/about/&#10;https://example.com/contact/" oninput="checkUrlLimit(this)"></textarea>
    <span id="url-count-msg" style="font-size:10px;color:#444"></span>
  </div>
  <div class="fg" id="t-sitemap" style="display:none">
    <label>Sitemap URL</label>
    <input type="text" id="sitemapInput" placeholder="https://example.com/sitemap.xml">
    <span style="font-size:10px;color:#444">First 10 URLs from sitemap will be audited</span>
  </div>
  <div class="fg">
    <label>Max URLs</label>
    <div style="background:#111;border:1px solid #1a1a1a;border-radius:7px;padding:8px 12px;font-size:12px;color:#22c55e;font-family:'DM Mono',monospace;display:flex;justify-content:space-between;align-items:center">
      <span>10 URLs per audit</span>
      <span style="font-size:10px;color:#444">Free plan limit</span>
    </div>
    <input type="hidden" id="maxUrls" value="10">
  </div>

  <button class="run-btn" id="runBtn" onclick="startAudit()">Run Audit →</button>
  <span class="hint">Crawl + redirect check + broken link scan runs automatically.</span>

  <div style="border-top:1px solid #111;padding-top:16px">
    <div class="section-h">Recent Audits</div>
    {history_rows}
  </div>
</aside>

<div class="panel" id="panel">
  <div class="idle" id="idle">
    <div class="idle-ico">🔬</div>
    <div class="idle-h">Ready to Audit</div>
    <div class="idle-sub">Enter a site name, paste URLs or a sitemap URL, and hit Run Audit.</div>
  </div>

  <div id="prog" style="display:none">
    <div class="prog-hdr"><span class="prog-t" id="progT">Auditing...</span><span class="prog-c" id="progC">0/0</span></div>
    <div class="prog-bg"><div class="prog-fill" id="progF" style="width:0%"></div></div>
    <div class="cur-url" id="curU"></div>
    <span class="badge b-crawl" id="phBadge">crawling pages</span>
  </div>

  <div id="rtabs" style="display:none">
    <div class="rtab-bar">
      <button class="rtab on"  onclick="showTab('overview',this)">Overview</button>
      <button class="rtab"     onclick="showTab('redirects',this)">Redirects</button>
      <button class="rtab"     onclick="showTab('links',this)">Broken Links</button>
      <button class="rtab"     onclick="showTab('images',this)">Images</button>
      <button class="rtab"     onclick="showTab('speed',this)">Page Speed</button>
    </div>
  </div>

  <!-- OVERVIEW -->
  <div class="rtab-c on" id="rt-overview">
    <div class="section-h">Overview</div>
    <div class="grid4" style="margin-bottom:14px">
      <div class="card neu"><div class="card-n" id="s-total">—</div><div class="card-l">Total URLs</div></div>
      <div class="card red"><div class="card-n" id="s-issues">—</div><div class="card-l">With Issues</div></div>
      <div class="card ora"><div class="card-n" id="s-errors">—</div><div class="card-l">Fetch Errors</div></div>
      <div class="card red"><div class="card-n" id="s-noindex">—</div><div class="card-l">Noindex</div></div>
    </div>
    <div class="section-h">Tag Coverage</div>
    <div class="grid2" style="margin-bottom:14px">
      <div class="irow"><span class="irow-n">Missing Title</span><span class="ib r" id="b-title">—</span></div>
      <div class="irow"><span class="irow-n">Missing Meta Desc</span><span class="ib r" id="b-desc">—</span></div>
      <div class="irow"><span class="irow-n">Missing H1</span><span class="ib r" id="b-h1">—</span></div>
      <div class="irow"><span class="irow-n">Missing Canonical</span><span class="ib o" id="b-can">—</span></div>
      <div class="irow"><span class="irow-n">No Structured Data</span><span class="ib o" id="b-schema">—</span></div>
      <div class="irow"><span class="irow-n">Slow Pages (&gt;2s)</span><span class="ib o" id="b-slow">—</span></div>
    </div>
    <div class="section-h">Crawl Log</div>
    <div class="log" id="logBox"></div>
  </div>

  <!-- REDIRECTS -->
  <div class="rtab-c" id="rt-redirects">
    <div class="grid3" style="margin-bottom:14px">
      <div class="card ora"><div class="card-n" id="s-redir">—</div><div class="card-l">Pages Redirecting</div></div>
      <div class="card neu"><div class="card-n" id="s-total2">—</div><div class="card-l">Total Crawled</div></div>
      <div class="card grn"><div class="card-n" id="s-noredir">—</div><div class="card-l">No Redirect</div></div>
    </div>
    <div class="section-h">Redirect Log</div>
    <div class="log" id="redirLog"></div>
  </div>

  <!-- BROKEN LINKS -->
  <div class="rtab-c" id="rt-links">
    <div id="lp-wrap" style="display:none;margin-bottom:14px">
      <div class="prog-hdr"><span class="prog-t" style="font-size:12px">Checking links...</span><span class="prog-c" id="lp-c">0/0</span></div>
      <div class="prog-bg"><div class="prog-fill" id="lp-f" style="width:0%"></div></div>
    </div>
    <div class="grid3" style="margin-bottom:14px">
      <div class="card red"><div class="card-n" id="s-broken">—</div><div class="card-l">Broken Links</div></div>
      <div class="card ora"><div class="card-n" id="s-bint">—</div><div class="card-l">Internal Broken</div></div>
      <div class="card ora"><div class="card-n" id="s-bext">—</div><div class="card-l">External Broken</div></div>
    </div>
    <div class="section-h">Broken Links</div>
    <div class="log" id="brokenLog"></div>
  </div>

  <!-- IMAGES -->
  <div class="rtab-c" id="rt-images">
    <div class="grid3" style="margin-bottom:14px">
      <div class="card neu"><div class="card-n" id="s-imgs">—</div><div class="card-l">Total Images</div></div>
      <div class="card red"><div class="card-n" id="s-imgiss">—</div><div class="card-l">Alt Issues</div></div>
      <div class="card grn"><div class="card-n" id="s-imgok">—</div><div class="card-l">Images OK</div></div>
    </div>
    <div class="section-h">Image Issues</div>
    <div class="log" id="imgLog"></div>
  </div>

  <!-- PAGE SPEED -->
  <div class="rtab-c" id="rt-speed">
    <div style="display:flex;align-items:center;gap:14px;margin-bottom:16px;flex-wrap:wrap">
      <button class="speed-btn" id="spdBtn" onclick="runSpeed()">▶ Run Speed Audit</button>
      <span style="font-size:11px;color:#333">Checks mobile Core Web Vitals via Google PageSpeed Insights.</span>
    </div>
    <div id="spd-prog" style="display:none;margin-bottom:14px">
      <div class="prog-hdr"><span class="prog-t" style="font-size:12px">Running PageSpeed...</span><span class="prog-c" id="sp-c"></span></div>
      <div class="prog-bg"><div class="prog-fill" id="sp-f" style="width:50%;animation:pulse 1s infinite"></div></div>
    </div>
    <div id="spd-err" style="display:none;font-size:12px;color:#f87171;padding:8px;background:#1a0808;border-radius:6px;margin-bottom:12px"></div>
    <div id="spd-cards" style="display:flex;flex-direction:column;gap:12px"></div>
    <div id="spd-save-row" style="display:none;margin-top:16px;padding-top:14px;border-top:1px solid #111">
      <button class="speed-btn" id="spdSaveBtn" onclick="saveSpeedToExcel()" style="background:#22c55e;color:#0a0a0a;border-color:#22c55e">⬇ Save Speed Data to Excel</button>
      <span id="spd-save-msg" style="font-size:11px;color:#444;margin-left:12px"></span>
    </div>
  </div>

  <!-- DOWNLOAD -->
  <div id="dl-sec" style="display:none;border-top:1px solid #111;padding-top:16px">
    <div class="section-h">Export Report</div>
    <div style="display:flex;align-items:center;gap:12px">
      <button class="dl-btn" id="dlBtn" onclick="dl()" disabled>⬇ Download Excel Report</button>
      <span class="done-tag" id="doneTag" style="display:none">✓ AUDIT COMPLETE — 7 SHEETS</span>
    </div>
  </div>
</div>
</main>

<script>
let tab='urls',jobId=null,pollT=null;
function checkUrlLimit(ta){{
  const urls=ta.value.split('\n').filter(u=>u.trim().startsWith('http'));
  const msg=document.getElementById('url-count-msg');
  if(urls.length>10){{
    msg.textContent=`⚠ Only first 10 of ${{urls.length}} URLs will be audited`;
    msg.style.color='#fb923c';
  }}else if(urls.length>0){{
    msg.textContent=`${{urls.length}} URL${{urls.length>1?'s':''}} entered`;
    msg.style.color='#22c55e';
  }}else{{
    msg.textContent='';
  }}
}}

function switchTab(t){{
  tab=t;
  // Show/hide input sections
  document.getElementById('t-urls').style.display=t==='urls'?'flex':'none';
  document.getElementById('t-sitemap').style.display=t==='sitemap'?'flex':'none';
  // Style the buttons directly
  const urlBtn=document.getElementById('btn-urls');
  const sitemapBtn=document.getElementById('btn-sitemap');
  if(t==='urls'){{
    urlBtn.style.background='#22c55e'; urlBtn.style.color='#0a0a0a'; urlBtn.style.fontWeight='700';
    sitemapBtn.style.background='transparent'; sitemapBtn.style.color='#555'; sitemapBtn.style.fontWeight='500';
  }}else{{
    sitemapBtn.style.background='#22c55e'; sitemapBtn.style.color='#0a0a0a'; sitemapBtn.style.fontWeight='700';
    urlBtn.style.background='transparent'; urlBtn.style.color='#555'; urlBtn.style.fontWeight='500';
  }}
  document.getElementById('url-count-msg').textContent='';
}}

function showTab(id,btn){{
  document.querySelectorAll('.rtab-c').forEach(e=>e.classList.remove('on'));
  document.querySelectorAll('.rtab').forEach(b=>b.classList.remove('on'));
  document.getElementById('rt-'+id).classList.add('on');
  btn.classList.add('on');
}}
async function startAudit(){{
  const siteName=document.getElementById('siteName').value.trim()||'Website';
  const maxUrls=parseInt(document.getElementById('maxUrls').value);
  let body={{site_name:siteName,max_urls:maxUrls}};
  if(tab==='sitemap'){{
    const u=document.getElementById('sitemapInput').value.trim();
    if(!u)return alert('Enter a sitemap URL.');
    body.sitemap_url=u;
  }}else{{
    const raw=document.getElementById('urlInput').value.trim();
    if(!raw)return alert('Paste at least one URL.');
    body.urls=raw;
  }}
  document.getElementById('runBtn').disabled=true;
  document.getElementById('idle').style.display='none';
  ['prog','rtabs','dl-sec'].forEach(id=>document.getElementById(id).style.display='block');
  document.getElementById('dlBtn').disabled=true;
  document.getElementById('doneTag').style.display='none';
  document.getElementById('logBox').innerHTML='';
  document.getElementById('brokenLog').innerHTML='';
  document.getElementById('imgLog').innerHTML='';
  const res=await fetch('/api/start',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify(body)}});
  const data=await res.json();
  if(data.error){{alert(data.error);document.getElementById('runBtn').disabled=false;return;}}
  jobId=data.job_id;
  document.getElementById('progT').textContent='Auditing '+siteName;
  pollT=setInterval(doPoll,1800);
}}
async function doPoll(){{
  if(!jobId)return;
  const res=await fetch('/api/status/'+jobId);
  const d=await res.json();
  const s=d.summary||{{}};
  const pct=d.total>0?Math.round(d.progress/d.total*100):0;
  document.getElementById('progF').style.width=pct+'%';
  document.getElementById('progC').textContent=d.progress+'/'+d.total;
  document.getElementById('curU').textContent=d.current_url||'';
  const pb=document.getElementById('phBadge');
  if(d.link_status==='running'){{
    pb.className='badge b-links';pb.textContent='checking links ('+d.link_progress+'/'+d.link_total+')';
    document.getElementById('lp-wrap').style.display='block';
    const lp=d.link_total>0?Math.round(d.link_progress/d.link_total*100):0;
    document.getElementById('lp-f').style.width=lp+'%';
    document.getElementById('lp-c').textContent=d.link_progress+'/'+d.link_total;
  }}else if(d.status==='done'){{
    pb.className='badge b-done';pb.textContent='complete';
    document.getElementById('lp-wrap').style.display='none';
  }}else{{pb.className='badge b-crawl';pb.textContent='crawling pages';}}
  if(s.total){{
    const set=(id,v)=>{{const e=document.getElementById(id);if(e)e.textContent=v;}};
    set('s-total',s.total);set('s-issues',s.issues);set('s-errors',s.fetch_errors);set('s-noindex',s.noindex);
    set('b-title',s.missing_title);set('b-desc',s.missing_desc);set('b-h1',s.missing_h1);
    set('b-can',s.missing_canonical);set('b-schema',s.no_schema);set('b-slow',s.slow_pages);
    set('s-redir',s.redirects);set('s-total2',s.total);set('s-noredir',s.total-s.redirects);
    set('s-broken',s.broken_links||0);set('s-bint',s.internal_broken||0);
    set('s-bext',(s.broken_links||0)-(s.internal_broken||0));
    set('s-imgs',s.total_images||0);set('s-imgiss',s.img_issues||0);
    set('s-imgok',(s.total_images||0)-(s.img_issues||0));
  }}
  const lb=document.getElementById('logBox');
  lb.innerHTML=(d.log||[]).slice().reverse().map(r=>{{
    const ok=r.status>=200&&r.status<300,rdr=r.status>=300&&r.status<400;
    return `<div class="lrow"><span class="ls ${{ok?'ok':rdr?'rdr':'err'}}">${{r.status||'ERR'}}</span><span class="lms">${{r.ms}}ms</span><span class="lu">${{r.url}}</span>${{r.issues>0?`<span class="li">${{r.issues}}⚠</span>`:''}}</div>`;
  }}).join('');
  const rl=document.getElementById('redirLog');
  const rdrs=(d.log||[]).filter(r=>r.status>=300&&r.status<400);
  if(rdrs.length)rl.innerHTML=rdrs.slice().reverse().map(r=>`<div class="lrow"><span class="ls rdr">${{r.status}}</span><span class="lu">${{r.url}}</span></div>`).join('');
  if(d.link_status==='done'&&s.broken_links>0&&!document.getElementById('brokenLog').innerHTML)
    document.getElementById('brokenLog').innerHTML=`<div class="lrow" style="color:#444;font-size:11px">Found ${{s.broken_links}} broken links (${{s.internal_broken}} internal). Full list in Excel → Broken Links sheet.</div>`;
  if(s.img_issues>0&&!document.getElementById('imgLog').innerHTML)
    document.getElementById('imgLog').innerHTML=`<div class="lrow" style="color:#444;font-size:11px">Found ${{s.img_issues}} images with alt issues across ${{s.total_images}} total. Full breakdown in Excel → Image Audit sheet.</div>`;
  if(d.speed_status==='running'){{
    document.getElementById('spd-prog').style.display='block';
    const sp=d.speed_total>0?Math.round(d.speed_progress/d.speed_total*100):0;
    document.getElementById('sp-f').style.width=sp+'%';
    document.getElementById('sp-c').textContent=d.speed_progress+'/'+d.speed_total;
  }}else if(d.speed_status==='done'){{
    document.getElementById('spd-prog').style.display='none';
    document.getElementById('spd-done').style.display='block';
    document.getElementById('spdBtn').disabled=false;
  }}
  if(d.status==='done'){{
    clearInterval(pollT);
    document.getElementById('progF').style.width='100%';
    document.getElementById('progC').textContent=d.total+'/'+d.total;
    document.getElementById('runBtn').disabled=false;
    if(d.has_excel){{document.getElementById('dlBtn').disabled=false;document.getElementById('doneTag').style.display='inline-block';}}
  }}
}}
let speedResultsCache={{}};

async function runSpeed(){{
  if(!jobId)return alert('Run an audit first.');
  speedResultsCache={{}};
  document.getElementById('spd-save-row').style.display='none';
  const btn=document.getElementById('spdBtn');
  btn.disabled=true;
  document.getElementById('spd-prog').style.display='block';
  document.getElementById('spd-err').style.display='none';
  document.getElementById('spd-cards').innerHTML='';
  document.getElementById('sp-c').textContent='Calling Google API...';

  // Get URLs from current audit results
  const statusRes=await fetch('/api/status/'+jobId);
  const statusData=await statusRes.json();
  const auditedUrls=(statusData.log||[]).map(r=>r.url).filter(u=>u).slice(0,5);

  if(auditedUrls.length===0){{
    document.getElementById('spd-err').textContent='No URLs found from current audit. Please run an audit first.';
    document.getElementById('spd-err').style.display='block';
    document.getElementById('spd-prog').style.display='none';
    btn.disabled=false; return;
  }}

  try{{
    const res=await fetch('/api/speed/direct',{{
      method:'POST',
      headers:{{'Content-Type':'application/json'}},
      body:JSON.stringify({{urls:auditedUrls}})
    }});
    const data=await res.json();
    document.getElementById('spd-prog').style.display='none';
    btn.disabled=false;
    if(data.error){{
      document.getElementById('spd-err').textContent='Error: '+data.error;
      document.getElementById('spd-err').style.display='block';
      return;
    }}
    // Render cards
    const cards=document.getElementById('spd-cards');
    cards.innerHTML='';
    for(const[url,sp] of Object.entries(data.results||{{}})){{
      const score=sp.perf_score;
      const sc=typeof score==='number'?score:'—';
      const scClass=typeof sc==='number'?(sc>=90?'grn':sc>=50?'ora':'red'):'neu';
      const badge=(v,label)=>{{
        const cls=v==='Good'?'bg':v==='Needs Improvement'?'bo':'br';
        return v&&v!=='—'?`<span class="badge ${{cls}}">${{v}}</span>`:`<span style="color:#444">${{v||'—'}}</span>`;
      }};
      cards.innerHTML+=`
        <div style="background:#0f0f0f;border:1px solid #1a1a1a;border-radius:9px;padding:16px">
          <div style="font-size:10px;color:#444;font-family:'DM Mono',monospace;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;margin-bottom:10px">${{url}}</div>
          ${{sp.error?`<div style="color:#f87171;font-size:12px">API Error: ${{sp.error}}</div>`:`
          <div style="display:flex;align-items:center;gap:16px;margin-bottom:12px">
            <div>
              <div style="font-size:11px;color:#444;text-transform:uppercase;letter-spacing:.08em;margin-bottom:2px">Performance</div>
              <div style="font-size:32px;font-weight:800;font-family:'DM Mono',monospace;color:${{typeof sc==='number'&&sc>=90?'#22c55e':sc>=50?'#fb923c':'#f87171'}}">${{sc}}</div>
            </div>
            <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;flex:1">
              <div style="background:#111;border-radius:6px;padding:8px">
                <div style="font-size:11px;font-family:'DM Mono',monospace;font-weight:500;margin-bottom:3px">${{sp.lcp||'—'}}</div>
                <div style="font-size:9px;color:#444">LCP ${{badge(sp.lcp_score)}}</div>
              </div>
              <div style="background:#111;border-radius:6px;padding:8px">
                <div style="font-size:11px;font-family:'DM Mono',monospace;font-weight:500;margin-bottom:3px">${{sp.cls||'—'}}</div>
                <div style="font-size:9px;color:#444">CLS ${{badge(sp.cls_score)}}</div>
              </div>
              <div style="background:#111;border-radius:6px;padding:8px">
                <div style="font-size:11px;font-family:'DM Mono',monospace;font-weight:500;margin-bottom:3px">${{sp.tbt||'—'}}</div>
                <div style="font-size:9px;color:#444">TBT ${{badge(sp.tbt_score)}}</div>
              </div>
              <div style="background:#111;border-radius:6px;padding:8px">
                <div style="font-size:11px;font-family:'DM Mono',monospace;font-weight:500;margin-bottom:3px">${{sp.fcp||'—'}}</div>
                <div style="font-size:9px;color:#444">FCP</div>
              </div>
              <div style="background:#111;border-radius:6px;padding:8px">
                <div style="font-size:11px;font-family:'DM Mono',monospace;font-weight:500;margin-bottom:3px">${{sp.ttfb||'—'}}</div>
                <div style="font-size:9px;color:#444">TTFB</div>
              </div>
            </div>
          </div>`}}
        </div>`;
    }}
  }}catch(e){{
    document.getElementById('spd-err').textContent='Request failed: '+e.message;
    document.getElementById('spd-err').style.display='block';
    document.getElementById('spd-prog').style.display='none';
    btn.disabled=false;
  }}
}}
function dl(){{if(jobId)window.location.href='/api/download/'+jobId;}}
</script>
</body></html>"""

# ══════════════════════════════════════════════════════════════════════
# MAIN — init_db runs here so gunicorn triggers it too
# ══════════════════════════════════════════════════════════════════════
with app.app_context():
    try:
        db = get_db()
        if USE_PG:
            cur = db.cursor()
            cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id       SERIAL PRIMARY KEY,
                    name     TEXT NOT NULL,
                    email    TEXT NOT NULL UNIQUE,
                    password TEXT NOT NULL,
                    plan     TEXT DEFAULT 'free',
                    created  TIMESTAMP DEFAULT NOW()
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS audits (
                    id         SERIAL PRIMARY KEY,
                    user_id    INTEGER NOT NULL,
                    site_name  TEXT,
                    url_count  INTEGER,
                    issues     INTEGER,
                    created    TIMESTAMP DEFAULT NOW()
                )
            """)
            db.commit()
            print("PostgreSQL tables ready")
        else:
            cur = db.cursor()
            cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id       INTEGER PRIMARY KEY AUTOINCREMENT,
                    name     TEXT NOT NULL,
                    email    TEXT NOT NULL UNIQUE,
                    password TEXT NOT NULL,
                    plan     TEXT DEFAULT 'free',
                    created  TEXT DEFAULT (datetime('now'))
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS audits (
                    id         INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id    INTEGER NOT NULL,
                    site_name  TEXT,
                    url_count  INTEGER,
                    issues     INTEGER,
                    created    TEXT DEFAULT (datetime('now'))
                )
            """)
            db.commit()
            print("SQLite tables ready")
    except Exception as e:
        print(f"DB init error: {e}")

if __name__ == "__main__":
    print("\n" + "="*54)
    print("  GetYouRank SEO Auditor")
    print("  Open: http://localhost:5000")
    print("="*54 + "\n")
    app.run(debug=False, port=5000, host="0.0.0.0")
